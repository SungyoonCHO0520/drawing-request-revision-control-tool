from __future__ import annotations

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .data_models import ALARM_LEVEL_OPTIONS, RESULT_OPTIONS, STATUS_OPTIONS, YES_NO_OPTIONS


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
PASS_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
NG_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
CHECK_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
MISSING_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
HIGH_FILL = PatternFill(fill_type="solid", fgColor="C00000")
MEDIUM_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
LOW_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
WHITE_BOLD = Font(color="FFFFFF", bold=True)


def header_map(ws) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in ws[1] if cell.value}


def autosize_columns(wb) -> None:
    for ws in wb.worksheets:
        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                max_length = max(max_length, len(str(cell.value or "")))
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 10), 48)


def apply_base_styles(wb) -> None:
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True, color="1F2933")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize_columns(wb)


def _add_formula(ws, target_range: str, formula: str, fill: PatternFill, font: Font | None = None) -> None:
    ws.conditional_formatting.add(target_range, FormulaRule(formula=[formula], fill=fill, font=font))


def apply_conditional_formatting(wb) -> None:
    for ws in wb.worksheets:
        headers = header_map(ws)
        last_row = max(ws.max_row, 2)
        if "Result" in headers:
            col = get_column_letter(headers["Result"])
            rng = f"{col}2:{col}{last_row}"
            _add_formula(ws, rng, f'{col}2="PASS"', PASS_FILL)
            _add_formula(ws, rng, f'{col}2="NG"', NG_FILL, Font(bold=True))
            _add_formula(ws, rng, f'{col}2="CHECK"', CHECK_FILL)
            _add_formula(ws, rng, f'{col}2="MISSING"', MISSING_FILL)
        if "Alarm Level" in headers:
            col = get_column_letter(headers["Alarm Level"])
            rng = f"{col}2:{col}{last_row}"
            _add_formula(ws, rng, f'{col}2="High"', HIGH_FILL, WHITE_BOLD)
            _add_formula(ws, rng, f'{col}2="Medium"', MEDIUM_FILL)
            _add_formula(ws, rng, f'{col}2="Low"', LOW_FILL)


def _validation(values: list[str]) -> DataValidation:
    return DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)


def apply_data_validation(wb) -> None:
    validations = {
        "Critical": _validation(YES_NO_OPTIONS),
        "Required": _validation(YES_NO_OPTIONS),
        "Confirmed": _validation(YES_NO_OPTIONS),
        "Result": _validation(RESULT_OPTIONS),
        "Alarm Level": _validation(ALARM_LEVEL_OPTIONS),
        "Status": _validation(STATUS_OPTIONS),
        "Change Risk": _validation(ALARM_LEVEL_OPTIONS),
    }
    for ws in wb.worksheets:
        headers = header_map(ws)
        last_row = max(ws.max_row + 200, 202)
        for column, validation in validations.items():
            if column in headers:
                copied = _validation(validation.formula1.strip('"').split(","))
                ws.add_data_validation(copied)
                col = get_column_letter(headers[column])
                copied.add(f"{col}2:{col}{last_row}")
