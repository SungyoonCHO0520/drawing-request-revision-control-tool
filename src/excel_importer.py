from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .data_models import SHEET_TO_TABLE, TABLE_SCHEMAS, normalize_dataframe
from .database import create_project, save_project


def _sheet_to_dataframe(ws) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    headers = [str(value or "").strip() for value in rows[0]]
    data = []
    for row in rows[1:]:
        if row is None:
            continue
        if not any(value not in (None, "") for value in row):
            continue
        data.append({headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))})
    return pd.DataFrame(data, columns=headers)


def import_excel_to_dataframes(input_path: str | Path) -> tuple[dict[str, pd.DataFrame], list[str]]:
    workbook = load_workbook(input_path, data_only=False)
    dataframes: dict[str, pd.DataFrame] = {}
    warnings: list[str] = []
    for sheet_name in workbook.sheetnames:
        table_name = SHEET_TO_TABLE.get(sheet_name)
        if not table_name:
            continue
        df = _sheet_to_dataframe(workbook[sheet_name])
        missing = [column for column in TABLE_SCHEMAS[table_name] if column not in df.columns]
        if missing:
            warnings.append(f"{sheet_name}: missing required columns {missing}")
        dataframes[table_name] = normalize_dataframe(df, table_name)
    for table_name in TABLE_SCHEMAS:
        if table_name not in dataframes:
            warnings.append(f"{table_name}: sheet not found, created empty table")
            dataframes[table_name] = normalize_dataframe(None, table_name)
    return dataframes, warnings


def import_excel_to_project(input_path: str | Path, output_project: str | Path) -> list[str]:
    dataframes, warnings = import_excel_to_dataframes(input_path)
    create_project(output_project)
    save_project(output_project, dataframes)
    return warnings
