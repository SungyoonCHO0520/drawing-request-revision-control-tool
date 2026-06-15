from __future__ import annotations

from openpyxl import load_workbook

from src.database import create_project
from src.data_models import SHEET_NAMES
from src.excel_exporter import export_project_to_excel


def test_excel_export_sheet_and_style_rules(tmp_path):
    project = tmp_path / "sample.pfcproj"
    output = tmp_path / "sample.xlsx"
    create_project(project, sample=True)

    export_project_to_excel(project, output)
    workbook = load_workbook(output)

    assert workbook.sheetnames == list(SHEET_NAMES.values())
    ws = workbook["Measurement_Result_DB"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref.startswith("A1:")
    assert len(ws.conditional_formatting) > 0
    assert len(ws.data_validations.dataValidation) > 0
